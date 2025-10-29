import os
import csv
from typing import List, Dict


def append_row_to_excel(path: str, row: Dict[str, str], headers: List[str]):
    """
    Adiciona uma linha a um arquivo Excel (.xlsx) com cabeçalho fixo.
    Se openpyxl não estiver disponível, grava/concatena em CSV como fallback.
    """
    try:
        from openpyxl import Workbook, load_workbook
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            # Garante cabeçalho
            if ws.max_row == 0:
                ws.append(headers)
            else:
                first_row = [cell.value for cell in ws[1]]
                if first_row != headers:
                    # reescreve cabeçalho se divergente
                    ws.insert_rows(1)
                    for i, h in enumerate(headers, start=1):
                        ws.cell(row=1, column=i, value=h)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)

        ws.append([row.get(h, "") for h in headers])
        wb.save(path)
    except ImportError:
        # Fallback para CSV (compatível com Excel)
        csv_path = path[:-5] + ".csv" if path.lower().endswith(".xlsx") else path + ".csv"
        file_exists = os.path.exists(csv_path)
        with open(csv_path, mode="a", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
            if not file_exists:
                writer.writeheader()
            writer.writerow(row)
